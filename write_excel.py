"""
write_excel.py — Write styled Excel workbook with evidence tracing.

Sheet 1: Patient data with prototype styling + cell comments (hover for evidence)
Sheet 2: Evidence Map — evidence quotes in every cell for audit trail
"""

from __future__ import annotations

from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment


def write_styled_workbook(
    data_df: pd.DataFrame,
    evidence_df: pd.DataFrame,
    confidence_df: pd.DataFrame,
    template_path: str | Path,
    output_path: str | Path,
) -> None:
    """
    Write the output Excel workbook with styling, evidence comments, and evidence sheet.
    """
    template_path = Path(template_path)
    output_path = Path(output_path)

    workbook = load_workbook(template_path)
    worksheet = workbook[workbook.sheetnames[0]]

    header_row = 1
    data_row_start = 2
    num_cols = len(data_df.columns)

    # ── Capture template styles from row 2 ──
    template_row_height = worksheet.row_dimensions[data_row_start].height
    template_styles = []
    for col_idx in range(1, num_cols + 1):
        cell = worksheet.cell(data_row_start, col_idx)
        template_styles.append({
            "_style": copy(cell._style),
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "number_format": cell.number_format,
            "protection": copy(cell.protection),
        })

    # ── Clear existing data rows ──
    if worksheet.max_row >= data_row_start:
        worksheet.delete_rows(data_row_start, worksheet.max_row - data_row_start + 1)

    # ── Write data rows with styling and evidence comments ──
    worksheet.insert_rows(data_row_start, amount=len(data_df))

    for row_offset in range(len(data_df)):
        target_row = data_row_start + row_offset

        if template_row_height is not None:
            worksheet.row_dimensions[target_row].height = template_row_height

        for col_idx in range(1, num_cols + 1):
            cell = worksheet.cell(target_row, col_idx)
            value = data_df.iloc[row_offset, col_idx - 1]

            # Apply template styling
            style = template_styles[col_idx - 1]
            cell._style = copy(style["_style"])
            cell.font = copy(style["font"])
            cell.fill = copy(style["fill"])
            cell.border = copy(style["border"])
            cell.alignment = copy(style["alignment"])
            cell.number_format = style["number_format"]
            cell.protection = copy(style["protection"])

            # Write value (None for empty/NaN)
            if value is None or (isinstance(value, float) and pd.isna(value)) or value == "":
                cell.value = None
            else:
                cell.value = value

            # Add evidence comment if cell is populated
            if cell.value is not None:
                evidence_text = str(evidence_df.iloc[row_offset, col_idx - 1] or "")
                conf = str(confidence_df.iloc[row_offset, col_idx - 1] or "none")
                if evidence_text:
                    comment_text = (
                        f"[Confidence: {conf.upper()}]\n\n"
                        f"Source evidence:\n\"{evidence_text}\""
                    )
                    comment = Comment(comment_text, "Cloud9 AI Extractor")
                    comment.width = 400
                    comment.height = 200
                    cell.comment = comment

    # ── Create "Evidence Map" sheet ──
    ev_sheet = workbook.create_sheet("Evidence Map")

    # Write headers
    for col_idx in range(1, num_cols + 1):
        ev_sheet.cell(1, col_idx, data_df.columns[col_idx - 1])

    # Write evidence values
    for row_offset in range(len(evidence_df)):
        for col_idx in range(1, num_cols + 1):
            ev_value = str(evidence_df.iloc[row_offset, col_idx - 1] or "")
            conf = str(confidence_df.iloc[row_offset, col_idx - 1] or "none")
            if ev_value:
                ev_sheet.cell(row_offset + 2, col_idx, f"[{conf}] {ev_value}")

    # ── Save ──
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
